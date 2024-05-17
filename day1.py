import pandas as p
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


#INPUT OF LIST OF DICTIONARY
L=[]
n=int(input("enter the number of records"))
for i in range(n):
    d={}
    a=input("enter the id")
    b=input("enter the Quantity")
    d["ID"]=a
    d["Quantity"]=b
    L.append(d) 

#r=p.DataFrame(L)
#print(r)
def Find_QTY(ide,L):
    for i in L:
        if i["ID"]==ide:
            return i["Quantity"]

    
path="C:\\Users\\hp\\Desktop\\PythonDTU\\day1.xlsx"

A=load_workbook(path)
B=A['Sheet1']
t=p.read_excel(path,sheet_name='Sheet1',usecols=("ID","Quantity"))
size=len(t)
for i in range(2,size+2):
    y='B'+str(i)
    rr='A'+str(i)
    B[y]=Find_QTY(B[rr].value,L)
    print(B[rr])
A.save(path)


'''
t=p.read_excel(path,sheet_name='Sheet1',usecols=("ID","Quantity"))
print(t)
for r in dataframe_to_rows(t):
    if(len(r)==3):
        print("ro:",r[0])
        if( r[0]!=None):
            r[2]=Find_QTY(r[1],L)
            print(r[2])
            print(r)
print(t)
 '''       
        
        
'''
A=load_workbook(path)
B=A["Sheet1"]
for i in dataframe_to_rows(,index=False):
    print(i)
    B.append(i)
A.save(path)

'''
   
